import os
from ModuloCertificados import AuxFunc
import pandas as pd
import Global
import openpyxl

#======================================================================================================================
#                                             Funciones para Validaciones
#======================================================================================================================

################## Comprobar duplicados ##################
def Duplicados(Documentos):
    listaLimpia = []
    listaRepetidos = []
    for dni in Documentos:
        if dni not in listaLimpia:
            listaLimpia.append(dni)
        else:
            listaRepetidos.append(dni)

    return listaRepetidos
    # devuelve lista de decumentos duplicados

################### Comprobar Listas de Panel ####################

#---- Separar en listas de aprobados y reprobados de panel ----
def crearListasAR(listaPanel, lAprobados, lReprobados):
    for participante in listaPanel:
        dni = participante[0]
        apellido = participante[1]
        condicion = participante[2]
        if condicion == "APROBADO":
            lAprobados.append((dni, apellido))
        elif condicion == "REPROBADO":
            lReprobados.append((dni, apellido))

#---- Comprobar que todos los aprobados estén certificados ----
def AprobadosCertificados(Documentos, lAprobados, certificadosNoAprobados, aprobadosNoCertificados):
    
    for dni in Documentos:
        flag = False
        for aprobado in lAprobados:
            if dni != aprobado[0]:
                flag = True
            
        if flag:
            certificadosNoAprobados.append(aprobado)

    for aprobado in lAprobados:
        if aprobado[0] not in Documentos:
            aprobadosNoCertificados.append(aprobado)

#---- Comprobar certificados certificados que no aparecen en la lista de aprobados ----
def comprobarRestoCertificados(certificadosNoAprobados, lReprobados, certificadosNoInscriptos):
    if len(certificadosNoAprobados) != 0:
        for certificado in certificadosNoAprobados:
            if certificado not in lReprobados:
                certificadosNoInscriptos.append(certificado)

    for doc in certificadosNoAprobados:
        if doc not in lReprobados:
            certificadosNoAprobados.remove(doc)

#======================================================================================================================
#                                                           RUN
#======================================================================================================================

def ValidarCertificados():
    ######################### Crear Lista de Documentos ###########################

    # devuelve directorio del programa
    saved_path = os.getcwd()

    # Crea lista de nombres del directorio
    file_list = AuxFunc.nameList(saved_path + "/Certificados")

    dni = []

    for file in file_list:
        doc = file.split("_")[0]
        dni.append(doc)

    print("Se encontraron " + str(len(dni)) + " certificados. \n")

    ######################### Crear Listas de Aprobados y Reprobados de panel ###########################

    # Extrae los documentos y condiciones finales del excel de panel
    archivo = pd.read_excel("Listas/panel.xls")
    
    Documentos = []
    Apellido = []
    Condicion = []
    columnas = archivo.columns
    columnas = columnas.tolist()
    

    val = archivo[columnas[Global.columna_dni_panel]]
    for elemento in val:
        elemento = str(elemento)
        Documentos.append(elemento)
    
    val = archivo[columnas[Global.columna_apellido_panel]]
    for elemento in val:
        elemento = str(elemento)
        Apellido.append(elemento)

    val = archivo[columnas[Global.columna_condicion_panel]]
    for elemento in val:
        elemento = str(elemento)
        Condicion.append(elemento)

    # Crea una lista de tuplas de la forma (dni, condicion)
    ListaPanel = []
    i = 0
    while i != len(Documentos):
        tupla = (Documentos[i], Apellido[i],Condicion[i])
        ListaPanel.append(tupla)
        i += 1

    # separa la lista de panel en lista de aprobados y Reprobados
    lAprobados = []
    lReprobados = []

    crearListasAR(ListaPanel, lAprobados, lReprobados)

    print("En panel hay: ")
    print("Participantes: " + str(len(lAprobados) + len(lReprobados)))
    print("Aprobados: " + str(len(lAprobados)))
    print("Reprobados: " + str(len(lReprobados)))
    print("\n")

    ############################# Mostrar Repetidos #################################

    Repetidos = Duplicados(dni)

    if len(Repetidos) != 0:
        print("Hay " + str(len(Repetidos)) + " certificados reperidos:")
        print(Repetidos)
    elif len(Repetidos) == 0:
        print("No hay Certificados duplicados.")

    ###################################### Validación de panel ##########################################
    certificadosNoAprobados = []
    aprobadosNoCertificados = []

    AprobadosCertificados(dni, lAprobados, certificadosNoAprobados, aprobadosNoCertificados)

    # Participantes que aprobaron y faltan certificar
    if len(aprobadosNoCertificados) != 0:
        print("Hay " + str(
            len(aprobadosNoCertificados)) + " participantes aprobados en panel que no fueron certificados: ")
        print(aprobadosNoCertificados)

    certificadosNoInscriptos = []
    comprobarRestoCertificados(certificadosNoAprobados, lReprobados, certificadosNoInscriptos)

    # Certificados que no deben ser emitidos porque están reprobados
    if len(certificadosNoAprobados) != 0:
        print("Se emitieron " + str(
            len(certificadosNoAprobados)) + " certificados a participantes que aparecen Reprobados en panel: ")
        print(certificadosNoAprobados)

    # Certificados emitidos que no aparecen en panel
    if len(certificadosNoInscriptos) != 0:
        print("Se emitieron " + str(
            len(certificadosNoInscriptos)) + " certificados a participantes que no aparecen en panel: ")
        print(certificadosNoInscriptos)
    
    if (len(aprobadosNoCertificados) == 0) and (len(certificadosNoAprobados)==0) and (len(certificadosNoInscriptos)==0):
        print("\nNo hay errores de certificación!")
    else:
       #* ============================ Exportar errores ============================
        flagErrorExport = input("\nDesea exportar la lista de errores a un .xls? (s/n): ")
        if flagErrorExport == 's' or flagErrorExport == 'S':
            # crear documento de aprovados y reprobados
            wn = openpyxl.Workbook()
    
            #! Aprobados que faltan certificar
            hoja = wn.active
            hoja.title = "Aprobados no certificados"
            # print(f'Hoja activa: {hoja.title}')
    
            hoja.append(('DNI', 'Apellido'))
            for alumno in aprobadosNoCertificados:
                hoja.append(alumno)
    
            #! Certificados no aprobados
            hoja2 = wn.create_sheet("Certificados no aprobados")
            wn.active = hoja2
    
            hoja2.append(('DNI', 'Apellido'))
            for alumno in certificadosNoAprobados:
                hoja2.append(alumno)
            
            #! Certificados no inscriptos
            hoja3 = wn.create_sheet("Certificados no inscriptos")
            wn.active = hoja3
    
            hoja3.append(('DNI', 'Apellido'))
            for alumno in certificadosNoInscriptos:
                hoja3.append(alumno)
            
            wn.save('Listas/ErroresEnCertificacion.xlsx')
            
            #! Repetidos
            hoja4 = wn.create_sheet("Certificados Repetidos")
            wn.active = hoja4
    
            hoja4.append(('DNI', 'Apellido'))
            for alumno in Repetidos:
                aux = (alumno, "#")
                hoja4.append(aux)
            
            wn.save('Listas/ErroresEnCertificacion.xlsx')
    
    print("Todo listo!")
            