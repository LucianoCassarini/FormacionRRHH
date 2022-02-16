import pandas as pd
import openpyxl

#///////////////////Crea Lista de Hojar del archivo///////////////////////////
wb = openpyxl.load_workbook("Listas/drive.xlsx")

#Lista de nombres de las hojas del documento
# hojas = wb.get_sheet_names()
hojas = wb.sheetnames

#====================================================================================================================
#                                       Listado de Aprobados/Reprobados
#====================================================================================================================
listasDni = []
listasCondicion = []
listasNombres = []

for hoja in hojas:
    # Extrae los documentos y condiciones finales del excel de panel
    archivo = pd.read_excel("Listas/drive.xlsx", sheet_name=hoja, header=3)
    # El parametro header=3 denota cual es la cabecera, empieza a contar
    # las filas desde 0 por lo tanto la cabecera es 3(fila 4).

    # ///////////////////Lectura y listado de la primer Hoja/////////////////////
    lista = []
    Documentos = []
    Condicion = []
    NomCompleto = []
    columnas = archivo.columns
    columnas = columnas.tolist()
    valores = archivo.values

    # Crear lista de dni
    flagNanDni = True
    val = archivo[columnas[0]]
    for elemento in val:
        elemento = str(elemento)
        if elemento == 'nan':
            flagNanDni = False
        if flagNanDni == True:
            Documentos.append(elemento)

    # Crear Lista de Resultados
    flagNanCondicion = True
    val = archivo[columnas[17]]
    for elemento in val:
        elemento = str(elemento)
        if elemento == 'nan':
            flagNanCondicion = False
        if flagNanCondicion == True:
            Condicion.append(elemento)

    # Crear Lista de Nombre/Apellido

    flagNanNombre = True
    valApellido = archivo[columnas[1]]
    valName = archivo[columnas[2]]

    nanNameFlag = True
    n = 0
    while n < len(valApellido):
        nombre = str(valName[n])
        apellido = str(valApellido[n])
        if nombre == 'nan' or apellido == 'nan':
            nanNameFlag = False

        if nanNameFlag == True:
            completo = nombre + " " + apellido
            NomCompleto.append(completo)
        n += 1


    # //////////////////Borrar .0 de la lista de documentos/////////////////////
    DocumentosLimpios = []
    for doc in Documentos:
        doc = doc.split(".")[0]
        DocumentosLimpios.append(doc)
    DocumentosLimpios

    #listas de lista
    listasDni.append(DocumentosLimpios)
    listasCondicion.append(Condicion)
    listasNombres.append(NomCompleto)

#/////////////////// Separar datos de Aprobados y Reprobados /////////////////////////
Aprobados = []
Reprobados = []

# n empieza de 5 porque el documento usado para realizar las pruebas tiene 5 hojas con mal formato
n = 0
while n < len(listasDni):
    j = 0
    while j < len(listasDni[n]):
        if (listasCondicion[n])[j] == "APROBADO":
            aux = (listasDni[n])[j] , (listasNombres[n])[j] , "APROBADO" , hojas[n].split('- Comisión ')[1]
            Aprobados.append(aux)
        elif listasCondicion[n][j] == "REPROBADO":
            aux = (listasDni[n])[j] , (listasNombres[n])[j] , "REPROBADO" , hojas[n].split('- Comisión ')[1]
            Reprobados.append(aux)

        j += 1
    n += 1

# print(Aprobados)
# print(Reprobados)

#////////////// Crear Tabla con datos Anteriores ///////////////
#crear documento de aprovados y reprobados
wn = openpyxl.Workbook()

#configuro la hoja de aprobados
hoja = wn.active
hoja.title = "Aprobados"
# print(f'Hoja activa: {hoja.title}')

hoja.append(('DNI', 'Nombre', 'Condicion', 'Comisión'))
for alumno in Aprobados:
    hoja.append(alumno)

#Crea lista de reprobados
hoja2 = wn.create_sheet("Reprobados")
wn.active = hoja2

hoja2.append(('DNI', 'Nombre', 'Condicion', 'Comisión'))
for alumno in Reprobados:
    hoja2.append(alumno)

wn.save('DriveProcesado.xlsx')

