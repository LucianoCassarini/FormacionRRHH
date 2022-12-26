import openpyxl
from openpyxl import load_workbook
import Global
from time import sleep
from progress.bar import ChargingBar


def cuilToDNI(cuil):
    cuil = str(cuil)
    cuil = cuil.split('.')[0]
    dniAux = 0
    if (len(cuil) > 8):
        dni1 = cuil[2]
        dni2 = cuil[3]
        dni3 = cuil[4]
        dni4 = cuil[5]
        dni5 = cuil[6]
        dni6 = cuil[7]
        dni7 = cuil[8]
        dni8 = cuil[9]
        dniAux = dni1 + dni2 + dni3 + dni4 + dni5 + dni6 + dni7 + dni8
    elif len(cuil) <= 8:
        dniAux = cuil
    return dniAux


def colorCelda(celda, color):
    # this gives you Hexadecimal value of the color
    color_in_hex = celda.fill.start_color.index
    if color_in_hex == color:
        return True
    else:
        return False


def filtrarDrive():
    print("Procesando Drive...")
    # ///////////////////Crea Lista de Hojar del archivo///////////////////////////
    wb = load_workbook("Listas/drive.xlsx", data_only=True)

    # Lista de nombres de las hojas del documento
    # hojas = wb.get_sheet_names()
    hojas = wb.sheetnames

    bar1 = ChargingBar('', max=(len(hojas)+2))
    bar1.next()

    #! Eliminar Hoja Esqueleto
    if Global.nombre_hoja_esqueleto_drive in hojas:
        indice = hojas.index(Global.nombre_hoja_esqueleto_drive)
        hojas.pop(indice)
        
    #! Verificar error de separador
    flagSeparadores = True
    for hoja in hojas:
        if Global.separador_comision_drive not in hoja:
            print("\n Por favor revisá el separador de nombre de hoja y comision en el archivo Global para evitar errores")
            print("\n")
            flagSeparadores = False
        
    if flagSeparadores:
        # * ========== COLORES ===========
        rojo = Global.rojo_no_encontrados  # ! -> No encontrados
        gris = Global.gris_final_lista  # ? -> Fin de la lista
        # * ==============================

        #! ============= Pruebitas =============
        # sh = wb[hojas[0]]
        # print(sh.cell(row=83, column=1).value)
        # print(colorCelda((sh.cell(row=84, column=1)), Gris))
        #! =====================================

        Aprobados = []
        Reprobados = []

        for hoja in hojas:
            sh = wb[hoja]
            fila = 5
            while colorCelda(sh.cell(row=fila, column=1), gris) != True:
                if colorCelda(sh.cell(row=fila, column=1), rojo) != True:
                    documento = cuilToDNI(
                        sh.cell(row=fila, column=(Global.columna_dni_drive)+1).value)
                    if (sh.cell(row=fila, column=(Global.columna_condicion_drive)+1).value) == 'APROBADO':
                        aux = documento, (sh.cell(row=fila, column=(Global.columna_apellido_drive)+1).value), (sh.cell(
                            row=fila, column=(Global.columna_nombre_drive)+1).value), (sh.cell(row=fila, column=(Global.columna_correo_drive)+1).value), "APROBADO", (hoja.split(Global.separador_comision_drive)[1])
                        Aprobados.append(aux)
                    elif (sh.cell(row=fila, column=(Global.columna_condicion_drive)+1).value) == 'REPROBADO':
                        aux = documento, (sh.cell(row=fila, column=(Global.columna_apellido_drive)+1).value), (sh.cell(
                            row=fila, column=(Global.columna_nombre_drive)+1).value), (sh.cell(row=fila, column=(Global.columna_correo_drive)+1).value), "REPROBADO", (hoja.split(Global.separador_comision_drive)[1])
                        Reprobados.append(aux)
                fila += 1
            sleep(0.2)
            bar1.next()

        # * ====================================================================================
        # *                      Crear excel con Aprobados y Reprobados
        # * ====================================================================================
        # crear documento de aprovados y reprobados
        wn = openpyxl.Workbook()

        # configuro la hoja de aprobados
        hoja = wn.active
        hoja.title = "Aprobados"
        # print(f'Hoja activa: {hoja.title}')

        hoja.append(('DNI', 'Apellido', 'Nombre',
                    'Correo', 'Condicion', 'Comisión'))
        for alumno in Aprobados:
            hoja.append(alumno)

        bar1.next()
        sleep(0.2)

        # Crea lista de reprobados
        hoja2 = wn.create_sheet("Reprobados")
        wn.active = hoja2

        hoja2.append(('DNI', 'Apellido', 'Nombre',
                     'Correo', 'Condicion', 'Comisión'))
        for alumno in Reprobados:
            hoja2.append(alumno)

        bar1.next()
        sleep(0.2)

        wn.save('Listas/DriveProcesado.xlsx')
        print("\n")
        print("Drive procesado con exito! (Buscar en carpeta 'Listas')\n")


# filtrarDrive()
