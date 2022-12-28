import os
import pandas as pd
import Global
from time import sleep
from progress.bar import ChargingBar
import openpyxl
from openpyxl import load_workbook

#---- Separar en listas de aprobados y reprobados de panel ----
def crearListasAR(listaPanel, lAprobados, lReprobados):
    for participante in listaPanel:
        dni = participante[0]
        correo = participante[1]
        condicion = participante[2]
        if condicion == "APROBADO":
            lAprobados.append((dni, correo))
        elif condicion == "REPROBADO":
            lReprobados.append((dni, correo))


#------ Devuelve los elementos que no estan en ambas listas -------
def noEncontrados(lista1, lista2):
    listaNoEncontrados = []

    for participante in lista1:
        flag = False
        for participanteAux in lista2:
            if participanteAux[0] == participante[0]:
                flag = True
            elif participanteAux[1] == participante[1]:
                flag = True
        if flag == False:
            listaNoEncontrados.append(participante)

    return listaNoEncontrados

#---------- Devuelve la lista sin duplicados -----------
def eliminarDuplicados(lista):
    listaLimpia = []
    for elem in lista:
        if elem not in listaLimpia:
            listaLimpia.append(elem)
    return listaLimpia

def ambasListas(lista1, lista2):
    elementosCompartidos = []
    
    for elem in lista1:
        if elem in lista2:
            elementosCompartidos.append(elem)
    
    return elementosCompartidos


#======================================================================================================================
#                                                           RUN
#======================================================================================================================
def ValidarErroresDrivePanel():
    flagDocPanel = os.path.exists("Listas/panel.xls")
    flagDocDrive = os.path.exists("Listas/DriveProcesado.xlsx")
    
    if flagDocPanel and flagDocDrive:   
        print("Comparando listas de Drive y Panel...")
        # Extrae los documentos y condiciones finales del excel de panel
        archivo = pd.read_excel("Listas/panel.xls")

        bar1 = ChargingBar('', max=5)
        bar1.next()

        Documentos = []
        Correos = []
        Condicion = []
        columnas = archivo.columns
        columnas = columnas.tolist()

        val = archivo[columnas[Global.columna_dni_panel]]
        for elemento in val:
            elemento = str(elemento)
            Documentos.append(elemento)

        val = archivo[columnas[Global.columna_correo_panel]]
        for elemento in val:
            elemento = str(elemento)
            Correos.append(elemento)

        val = archivo[columnas[Global.columna_condicion_panel]]
        for elemento in val:
            elemento = str(elemento)
            Condicion.append(elemento)

        sleep(0.2)
        bar1.next()

        # Crea una lista de tuplas de la forma (dni, condicion)
        ListaPanel = []
        i = 0
        while i != len(Documentos):
            tupla = (Documentos[i], Correos[i], Condicion[i])
            ListaPanel.append(tupla)
            i += 1

        # separa la lista de panel en lista de aprobados y Reprobados
        AprobadosPanel = []
        ReprobadosPanel = []

        crearListasAR(ListaPanel, AprobadosPanel, ReprobadosPanel)

        sleep(0.2)
        bar1.next()

        # ================================COMPROBAR PANEL/DRIVE===================================
        #* ===========  Aprobados Drive ============
        AprobadosDrive = []
        ReprobadosDrive = []

        # Extrae los documentos y condiciones finales del excel de Drive
        archivo = pd.read_excel("Listas/DriveProcesado.xlsx", sheet_name=0, engine='openpyxl')
        dniAprobados = []
        correoAprobados = []
        AprobadosDrive = []
        columnas = archivo.columns
        columnas = columnas.tolist()

        val = archivo[columnas[0]]
        for elemento in val:
            elemento = str(elemento)
            dniAprobados.append(elemento)

        val = archivo[columnas[3]]
        for elemento in val:
            elemento = str(elemento)
            correoAprobados.append(elemento)

        i = 0
        while i != len(dniAprobados):
            tupla = (dniAprobados[i], correoAprobados[i])
            AprobadosDrive.append(tupla)
            i += 1

        sleep(0.2)
        bar1.next()

        #* ============ Reprobados Drive  ============
        archivo = pd.read_excel("Listas/DriveProcesado.xlsx", sheet_name=1, engine='openpyxl')
        dniReprobados = []
        correoReprobados = []
        columnas = archivo.columns
        columnas = columnas.tolist()

        val = archivo[columnas[0]]
        for elemento in val:
            elemento = str(elemento)
            dniReprobados.append(elemento)

        val = archivo[columnas[3]]
        for elemento in val:
            elemento = str(elemento)
            correoReprobados.append(elemento)

        i = 0
        while i != len(dniReprobados):
            tupla = (dniReprobados[i], correoReprobados[i])
            ReprobadosDrive.append(tupla)
            i += 1

        sleep(0.2)
        bar1.next()

        # ================Verificación================

        # # Aprobados que están reprobados en Panel
        # aNoPanel = noEncontrados(AprobadosDrive, AprobadosPanel)
        # # Aprobados que están reprobados en Drive
        # aNoDrive = noEncontrados(AprobadosPanel, AprobadosDrive)
        # # Reprobados que están aprobados en panel
        # rNoPanel = noEncontrados(ReprobadosDrive, ReprobadosPanel)
        # # Reprobados que están aprobados en panel
        # rNoDrive = noEncontrados(ReprobadosPanel, ReprobadosDrive)

        condicionARevisar = ambasListas(AprobadosDrive, ReprobadosPanel) + ambasListas(ReprobadosDrive, AprobadosPanel)
        condicionARevisar = eliminarDuplicados(condicionARevisar)

        panel = AprobadosPanel + ReprobadosPanel
        duplicadosPanel = ambasListas(AprobadosPanel, ReprobadosPanel)

        drive = AprobadosDrive + ReprobadosDrive
        duplicadosDrive = ambasListas(AprobadosDrive, ReprobadosDrive)


        faltantesPanel = noEncontrados(drive, panel)
        faltantesPanel = eliminarDuplicados(faltantesPanel)

        faltantesDrive = noEncontrados(panel, drive)
        faltantesDrive = eliminarDuplicados(faltantesDrive)



        # ======================================================================================================================
        #                                               Mostrar Resultados
        # ======================================================================================================================
        print("\n")

        # ------------ Comprobación de Aprobados -----------------
        flagNoError = True

        if len(duplicadosDrive) != 0:
            print('Los siguientes usuarios están duplicados en Drive:')
            print(duplicadosDrive)
            flagNoError = False

        if len(duplicadosPanel) != 0:
            print('Los siguientes usuarios están duplicados en Panel:')
            print(duplicadosPanel)
            flagNoError = False

        if len(faltantesPanel) != 0:
            print('Los siguientes usuarios no se encontraron en panel: ')
            print(faltantesPanel)
            flagNoError = False

        if len(faltantesDrive) != 0:
            print('Los siguientes usuarios no se encontraron en drive: ')
            print(faltantesDrive)
            flagNoError = False

        if len(condicionARevisar) != 0:
            print('Revisar la condición de los siguientes usuarios: ')
            print(condicionARevisar)
            flagNoError = False


        if flagNoError:
            print('No se encontraron errores!')
        else:
            #* ============================ Exportar errores ============================
            flagErrorExport = input("\nDesea exportar la lista de errores a un .xls? (s/n): ")
            if flagErrorExport == 's' or flagErrorExport == 'S':
                # crear documento de aprovados y reprobados
                wn = openpyxl.Workbook()

                # configuro la hoja no panel
                hoja = wn.active
                hoja.title = "No están en panel"
                # print(f'Hoja activa: {hoja.title}')

                hoja.append(('No están en panel', ' '))
                for alumno in faltantesPanel:
                    hoja.append(alumno)

                # Crea lista de no drive
                hoja2 = wn.create_sheet("No están en Drive")
                wn.active = hoja2

                hoja2.append(('No están en Drive', ' '))
                for alumno in faltantesDrive:
                    hoja2.append(alumno)

                # Crea lista de revisar condiciń
                hoja3 = wn.create_sheet("Revisar condición")
                wn.active = hoja3

                hoja3.append(('Revisar condición', ' '))
                for alumno in condicionARevisar:
                    hoja3.append(alumno)

                # Crea lista de duplicados panel
                hoja4 = wn.create_sheet("Duplicados panel")
                wn.active = hoja4

                hoja4.append(('Duplicados panel', ' '))
                for alumno in duplicadosPanel:
                    hoja4.append(alumno)

                # Crea lista de duplicados drive
                hoja5 = wn.create_sheet("Duplicados drive")
                wn.active = hoja5

                hoja5.append(('Duplicados panel', ' '))
                for alumno in duplicadosDrive:
                    hoja5.append(alumno)

                wn.save('Listas/ErroresDrivePanel.xlsx')

        print("Todo listo!")
    
    elif flagDocDrive == False:
        print("Primero debe ejecutar la acción [1].\n")
    else:
        print('No se encontro el archivo panel en la carpeta "Listas"')
        print('Por favor comprobar que el archivo se encuentre y el nombre del mismo sea "panel.xls" \n')